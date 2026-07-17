"""Tests unitarios del parser de estado de cuenta contra la fixture sintética."""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook

from portfoliosentinel.tools.exceptions import (
    MalformedStatementError,
    RowValidationError,
    TotalsMismatchError,
)
from portfoliosentinel.tools.parser import parse_account_statement
from tests.parser import expected as exp

REPO_ROOT = Path(__file__).resolve().parents[2]
FIXTURE = REPO_ROOT / "fixtures" / "estadocuenta-sintetico.xlsx"


@pytest.fixture(scope="module")
def snapshot():
    assert FIXTURE.is_file(), f"Falta la fixture {FIXTURE}"
    return parse_account_statement(FIXTURE)


def test_scrubbing_alias_only(snapshot):
    assert snapshot.investor_alias == exp.INVESTOR_ALIAS
    dumped = snapshot.model_dump_json()
    assert "Titular-Sintetico" not in dumped
    assert "COMITENTE-FICTICIO" not in dumped
    assert "INV-001" in dumped


def test_as_of(snapshot):
    assert snapshot.as_of == exp.AS_OF


def test_cash_exact(snapshot):
    by_ccy = {c.currency: c.amount for c in snapshot.cash}
    assert by_ccy["ARS"] == exp.CASH_ARS
    assert by_ccy["USD"] == exp.CASH_USD


def test_positions_exact_to_the_cent(snapshot):
    assert len(snapshot.positions) == len(exp.POSITIONS)
    for got, want in zip(snapshot.positions, exp.POSITIONS, strict=True):
        assert got.ticker == want["ticker"]
        assert got.quantity == want["quantity"]
        assert got.price == want["price"]
        assert got.total == want["total"]
        assert got.asset_class == want["asset_class"]
        assert got.quantity * got.price == got.total


def test_totals_and_mep_exact(snapshot):
    assert snapshot.total_ars == exp.TOTAL_ARS
    assert snapshot.total_usd == exp.TOTAL_USD
    assert snapshot.mep_implied == exp.MEP_IMPLIED
    assert snapshot.total_ars / snapshot.total_usd == snapshot.mep_implied


def test_overconcentration_planted(snapshot):
    ggal = next(p for p in snapshot.positions if p.ticker == "GGAL")
    weight = ggal.total / snapshot.total_ars
    assert weight == exp.GGAL_WEIGHT
    assert weight > Decimal("0.5")


def test_malformed_missing_section(tmp_path: Path):
    path = tmp_path / "bad_missing.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Titular", "Titular-Sintetico-Fixture"])
    ws.append(["MONEDAS"])
    ws.append(["Moneda", "Saldo"])
    ws.append(["ARS", 100.0])
    ws.append(["USD", 10.0])
    # Faltan ACCIONES/BONOS/CEDEARS/TOTALES
    wb.save(path)

    with pytest.raises(MalformedStatementError, match="Falta la sección obligatoria"):
        parse_account_statement(path)


def test_malformed_row_qty_price_mismatch(tmp_path: Path):
    path = tmp_path / "bad_row.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Titular", "Titular-Sintetico-Fixture"])
    ws.append(["Comitente", "COMITENTE-FICTICIO-000"])
    ws.append(["Fecha", "2026-07-15"])
    ws.append(["MONEDAS"])
    ws.append(["Moneda", "Saldo"])
    ws.append(["ARS", 0.0])
    ws.append(["USD", 0.0])
    ws.append(["ACCIONES"])
    ws.append(["Ticker", "Cantidad", "Precio", "Total"])
    ws.append(["GGAL", 10, 100.0, 999.0])  # 10*100 ≠ 999
    ws.append(["BONOS"])
    ws.append(["Ticker", "Cantidad", "Precio", "Total"])
    ws.append(["CEDEARS"])
    ws.append(["Ticker", "Cantidad", "Precio", "Total"])
    ws.append(["TOTALES"])
    ws.append(["Total ARS", 999.0])
    ws.append(["Total USD", 1.0])
    wb.save(path)

    with pytest.raises(RowValidationError, match="cantidad×precio"):
        parse_account_statement(path)


def test_malformed_totals_mismatch(tmp_path: Path):
    path = tmp_path / "bad_totals.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Titular", "Titular-Sintetico-Fixture"])
    ws.append(["Comitente", "COMITENTE-FICTICIO-000"])
    ws.append(["Fecha", "2026-07-15"])
    ws.append(["MONEDAS"])
    ws.append(["Moneda", "Saldo"])
    ws.append(["ARS", 1000.0])
    ws.append(["USD", 10.0])
    ws.append(["ACCIONES"])
    ws.append(["Ticker", "Cantidad", "Precio", "Total"])
    ws.append(["GGAL", 1, 500.0, 500.0])
    ws.append(["BONOS"])
    ws.append(["Ticker", "Cantidad", "Precio", "Total"])
    ws.append(["CEDEARS"])
    ws.append(["Ticker", "Cantidad", "Precio", "Total"])
    ws.append(["TOTALES"])
    ws.append(["Total ARS", 9999.0])  # debería ser 1500
    ws.append(["Total USD", 10.0])
    wb.save(path)

    with pytest.raises(TotalsMismatchError, match="Total ARS"):
        parse_account_statement(path)


def test_missing_file():
    with pytest.raises(MalformedStatementError, match="inexistente"):
        parse_account_statement("/tmp/no-existe-portfoliosentinel.xlsx")
