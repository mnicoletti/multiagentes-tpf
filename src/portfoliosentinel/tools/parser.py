"""Parser determinista de estado de cuenta .xlsx (ADR-0002 / ADR-0006 / ADR-0010).

Fuente de verdad numérica: openpyxl + Pydantic. Ningún LLM toca estos números.
Scrubbing de PII (titular/comitente → INV-001) ocurre antes de devolver el snapshot.

Soporta dos layouts:
- *compact* (fixture sintética): Ticker|Cantidad|Precio|Total + sección TOTALES;
  validación estricta qty×precio==total y montos a centavo.
- *broker_wide* (export típico de bróker AR): ESPECIE + CANT. DISPONIBLE + VALOR CORRIENTE;
  totales en cabecera; valor declarado del bróker como verdad de fila (ADR-0010).
"""

from __future__ import annotations

import logging
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook

from portfoliosentinel.config.settings import INVESTOR_ALIAS
from portfoliosentinel.tools.exceptions import (
    MalformedStatementError,
    RowValidationError,
    TotalsMismatchError,
)
from portfoliosentinel.tools.schemas import (
    AccountSnapshot,
    AssetClass,
    CashBalance,
    CurrencyCode,
    Position,
)

logger = logging.getLogger(__name__)

SECTION_MONEDAS = "MONEDAS"
SECTION_ACCIONES = "ACCIONES"
SECTION_BONOS = "BONOS"
SECTION_CEDEARS = "CEDEARS"
SECTION_TOTALES = "TOTALES"
ASSET_SECTIONS: tuple[AssetClass, ...] = ("ACCIONES", "BONOS", "CEDEARS")
ALL_SECTIONS = {SECTION_MONEDAS, *ASSET_SECTIONS, SECTION_TOTALES}

LayoutKind = Literal["compact", "broker_wide"]

# Tolerancia de comparación de totales en layout broker (ruido float de Excel).
_BROKER_TOTAL_TOLERANCE = Decimal("0.05")

_SKIP_ROW_LABELS = frozenset(
    {
        "subtotal",
        "moneda",
        "ticker",
        "especie",
        "instrumento",
        "concepto",
    }
)


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_decimal(value: Any, *, context: str) -> Decimal:
    if value is None or _cell_str(value) == "":
        raise MalformedStatementError(f"Valor numérico vacío en {context}")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        # Evitar binario flotante: pasar por str del valor que entrega openpyxl.
        return Decimal(str(value))
    try:
        return Decimal(str(value).strip().replace(",", ""))
    except (InvalidOperation, ValueError) as exc:
        raise MalformedStatementError(
            f"No se pudo interpretar número en {context}: {value!r}"
        ) from exc


def _parse_as_of(value: Any) -> date | None:
    if value is None or _cell_str(value) == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _cell_str(value)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise MalformedStatementError(f"Fecha de estado no reconocida: {value!r}")


def _quantize_cent(value: Decimal) -> Decimal:
    """Normaliza a centavos sin redondear montos intermedios: exige ya 2 decimales exactos."""
    normalized = value.quantize(Decimal("0.01"))
    if normalized != value:
        raise RowValidationError(f"Monto con más de 2 decimales (prohibido redondear): {value}")
    return normalized


def _read_sheet_rows(path: Path) -> list[tuple[Any, ...]]:
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001 — tipamos como malformado
        raise MalformedStatementError(f"No se pudo abrir el .xlsx '{path}': {exc}") from exc

    try:
        if not wb.sheetnames:
            raise MalformedStatementError("El .xlsx no tiene hojas")
        ws = wb[wb.sheetnames[0]]
        rows: list[tuple[Any, ...]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append(tuple(row))
        return rows
    finally:
        wb.close()


def _norm_header(text: str) -> str:
    return re.sub(r"\s+", " ", text.casefold().strip())


def _detect_layout(rows: list[tuple[Any, ...]]) -> LayoutKind:
    """Detecta compact (fixture) vs broker_wide (export de bróker con columnas anchas)."""
    for row in rows:
        cells = [_norm_header(_cell_str(c)) for c in row if _cell_str(c)]
        if not cells:
            continue
        joined = " | ".join(cells)
        if "especie" in cells and (
            "valor corriente" in cells or "cant. disponible" in cells or "cant disponible" in cells
        ):
            return "broker_wide"
        if "valor corriente" in joined and "moneda" in cells:
            return "broker_wide"
        # Cabecera típica de bróker
        if any("total cartera" in c for c in cells):
            return "broker_wide"
    return "compact"


def _find_header_fields_compact(
    rows: list[tuple[Any, ...]],
) -> tuple[date | None, str | None, str | None]:
    as_of: date | None = None
    titular: str | None = None
    comitente: str | None = None

    for row in rows:
        if not row:
            continue
        key = _cell_str(row[0]).casefold()
        val = row[1] if len(row) > 1 else None
        if key in {"fecha", "fecha de liquidación", "fecha liquidacion"}:
            as_of = _parse_as_of(val)
        elif key in {"titular", "cliente", "nombre"}:
            titular = _cell_str(val) or None
        elif key in {"comitente", "nro comitente", "número de comitente", "numero de comitente"}:
            comitente = _cell_str(val) or None
    return as_of, titular, comitente


def _find_header_fields_broker(
    rows: list[tuple[Any, ...]],
) -> tuple[date | None, str | None, str | None]:
    """Layout lateral: TITULAR | FECHA en una fila; valores en la siguiente; COMITENTE + fila."""
    as_of: date | None = None
    titular: str | None = None
    comitente: str | None = None

    for idx, row in enumerate(rows):
        labels = {_norm_header(_cell_str(c)): i for i, c in enumerate(row) if _cell_str(c)}
        if "titular" in labels or "cliente" in labels or "fecha" in labels:
            next_row = rows[idx + 1] if idx + 1 < len(rows) else ()
            if "titular" in labels or "cliente" in labels:
                col = labels.get("titular", labels.get("cliente", 0))
                if col < len(next_row):
                    titular = _cell_str(next_row[col]) or titular
            if "fecha" in labels:
                col = labels["fecha"]
                if col < len(next_row):
                    as_of = _parse_as_of(next_row[col]) or as_of
        key0 = _norm_header(_cell_str(row[0]) if row else "")
        if key0 in {"comitente", "nro comitente", "número de comitente", "numero de comitente"}:
            # Valor en col B de la misma fila, o sola en la siguiente.
            same = _cell_str(row[1]) if len(row) > 1 else ""
            if same:
                comitente = same
            elif idx + 1 < len(rows):
                comitente = _cell_str(rows[idx + 1][0]) or comitente

    # Fallback compact keys por si vienen mezclados
    if as_of is None or titular is None or comitente is None:
        a2, t2, c2 = _find_header_fields_compact(rows)
        as_of = as_of or a2
        titular = titular or t2
        comitente = comitente or c2
    return as_of, titular, comitente


def _find_header_totals_broker(
    rows: list[tuple[Any, ...]],
) -> tuple[Decimal | None, Decimal | None]:
    total_ars: Decimal | None = None
    total_usd: Decimal | None = None
    for row in rows:
        if not row:
            continue
        label = _norm_header(_cell_str(row[0]))
        if not label:
            continue
        # Primer valor numérico no vacío de la fila
        num: Any = None
        for cell in row[1:]:
            if cell is None or _cell_str(cell) == "":
                continue
            num = cell
            break
        if num is None:
            continue
        if "total cartera" in label or ("total" in label and "peso" in label):
            total_ars = _to_decimal(num, context="header/TOTAL_ARS")
        elif "total usd" in label or ("total" in label and "dolar" in label):
            total_usd = _to_decimal(num, context="header/TOTAL_USD")
    return total_ars, total_usd


def _iter_sections(
    rows: list[tuple[Any, ...]],
) -> dict[str, list[tuple[Any, ...]]]:
    """Parte la hoja en bloques por marcador de sección en columna A."""
    sections: dict[str, list[tuple[Any, ...]]] = {}
    current: str | None = None

    for row in rows:
        label = _cell_str(row[0]).upper() if row else ""
        if label in ALL_SECTIONS:
            current = label
            sections.setdefault(current, [])
            continue
        if current is not None:
            headerish = _norm_header(label) in _SKIP_ROW_LABELS
            if headerish and _norm_header(label) != "subtotal":
                # Conservamos la fila de encabezado de columnas para broker_wide
                # (especie/moneda) — se filtra después al mapear.
                if _norm_header(label) in {"especie", "moneda", "ticker", "instrumento"}:
                    sections[current].append(row)
                continue
            if _norm_header(label) == "subtotal":
                continue
            if all(_cell_str(c) == "" for c in row):
                continue
            sections[current].append(row)

    return sections


def _column_map(header_row: tuple[Any, ...]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for i, cell in enumerate(header_row):
        name = _norm_header(_cell_str(cell))
        if name:
            mapping[name] = i
    return mapping


def _col(mapping: dict[str, int], *candidates: str) -> int | None:
    for c in candidates:
        if c in mapping:
            return mapping[c]
    return None


def _parse_cash_compact(rows: list[tuple[Any, ...]]) -> list[CashBalance]:
    cash: list[CashBalance] = []
    seen: set[str] = set()
    for idx, row in enumerate(rows, start=1):
        currency_raw = _cell_str(row[0]).upper()
        if currency_raw in {"MONEDA", ""}:
            continue
        if currency_raw not in {"ARS", "USD"}:
            raise MalformedStatementError(
                f"MONEDAS fila {idx}: moneda desconocida '{currency_raw}' (esperado ARS|USD)"
            )
        currency: CurrencyCode = currency_raw  # type: ignore[assignment]
        if currency in seen:
            raise MalformedStatementError(f"MONEDAS: moneda duplicada '{currency}'")
        seen.add(currency)
        raw_amount = row[1] if len(row) > 1 else None
        amount = _quantize_cent(_to_decimal(raw_amount, context=f"MONEDAS/{currency}"))
        cash.append(CashBalance(currency=currency, amount=amount))
    if "ARS" not in seen or "USD" not in seen:
        raise MalformedStatementError("MONEDAS debe incluir saldos ARS y USD")
    return cash


def _parse_cash_broker(rows: list[tuple[Any, ...]]) -> tuple[list[CashBalance], Decimal]:
    """Agrega líneas USD; mapea $→ARS. Devuelve cash + suma de VALOR CORRIENTE (ARS)."""
    if not rows:
        raise MalformedStatementError("MONEDAS vacío")

    header_idx = 0
    mapping = _column_map(rows[0])
    if "moneda" not in mapping and "especie" not in mapping:
        # Sin header explícito: asumir col0 moneda, col2 cant, col4 valor
        mapping = {
            "moneda": 0,
            "cant. disponible": 2 if len(rows[0]) > 2 else 1,
            "valor corriente": 4 if len(rows[0]) > 4 else 2,
        }
        header_idx = -1
    else:
        header_idx = 0

    c_mon = _col(mapping, "moneda")
    c_qty = _col(mapping, "cant. disponible", "cant disponible", "cantidad", "saldo")
    c_val = _col(mapping, "valor corriente", "valor")
    if c_mon is None or c_qty is None:
        raise MalformedStatementError("MONEDAS broker: faltan columnas Moneda / Cant. disponible")

    ars_amount = Decimal("0")
    usd_amount = Decimal("0")
    valor_sum = Decimal("0")
    saw_ars = False
    saw_usd = False

    start = header_idx + 1
    for idx, row in enumerate(rows[start:], start=start + 1):
        label = _norm_header(_cell_str(row[c_mon] if c_mon < len(row) else ""))
        if not label or label == "subtotal":
            continue
        qty = _to_decimal(row[c_qty] if c_qty < len(row) else None, context=f"MONEDAS fila {idx}")
        if c_val is not None and c_val < len(row) and _cell_str(row[c_val]):
            valor_sum += _to_decimal(row[c_val], context=f"MONEDAS/valor fila {idx}")

        if label in {"$", "ars", "peso", "pesos"}:
            ars_amount += qty
            saw_ars = True
        elif label == "usd":
            usd_amount += qty
            saw_usd = True
        else:
            raise MalformedStatementError(
                f"MONEDAS fila {idx}: moneda desconocida '{label}' (esperado $|ARS|USD)"
            )

    if not saw_ars or not saw_usd:
        raise MalformedStatementError("MONEDAS debe incluir saldos ARS ($) y USD")

    cash = [
        CashBalance(currency="ARS", amount=ars_amount),
        CashBalance(currency="USD", amount=usd_amount),
    ]
    return cash, valor_sum


def _parse_positions_compact(
    rows: list[tuple[Any, ...]], asset_class: AssetClass
) -> list[Position]:
    positions: list[Position] = []
    for idx, row in enumerate(rows, start=1):
        label0 = _norm_header(_cell_str(row[0]) if row else "")
        if label0 in {"ticker", "especie", "instrumento", ""}:
            continue
        if len(row) < 4:
            raise MalformedStatementError(
                f"{asset_class} fila {idx}: se esperaban Ticker|Cantidad|Precio|Total"
            )
        ticker = _cell_str(row[0]).upper()
        if not ticker:
            raise MalformedStatementError(f"{asset_class} fila {idx}: ticker vacío")
        quantity = _to_decimal(row[1], context=f"{asset_class}/{ticker}/cantidad")
        price = _quantize_cent(_to_decimal(row[2], context=f"{asset_class}/{ticker}/precio"))
        total = _quantize_cent(_to_decimal(row[3], context=f"{asset_class}/{ticker}/total"))
        expected = quantity * price
        expected_cents = expected.quantize(Decimal("0.01"))
        if expected != total and expected_cents != total:
            raise RowValidationError(
                f"{asset_class} {ticker}: cantidad×precio={expected} ≠ total={total}"
            )
        if expected != total:
            raise RowValidationError(
                f"{asset_class} {ticker}: cantidad×precio={expected} ≠ total={total} "
                "(prohibido redondear)"
            )
        positions.append(
            Position(
                ticker=ticker,
                quantity=quantity,
                price=price,
                total=total,
                asset_class=asset_class,
            )
        )
    return positions


def _parse_positions_broker(rows: list[tuple[Any, ...]], asset_class: AssetClass) -> list[Position]:
    """Usa VALOR CORRIENTE como total declarado; no exige qty×precio==total (ADR-0010)."""
    if not rows:
        return []

    mapping = _column_map(rows[0])
    has_header = _col(mapping, "especie", "ticker", "instrumento") is not None
    if not has_header:
        mapping = {
            "especie": 0,
            "cant. disponible": 2,
            "precio": 4,
            "valor corriente": 6 if len(rows[0]) > 6 else 3,
        }
        data_rows = rows
    else:
        data_rows = rows[1:]

    c_ticker = _col(mapping, "especie", "ticker", "instrumento")
    c_qty = _col(mapping, "cant. disponible", "cant disponible", "cantidad")
    c_price = _col(mapping, "precio")
    c_total = _col(mapping, "valor corriente", "total", "valor moneda cotización")
    if c_ticker is None or c_qty is None or c_price is None or c_total is None:
        raise MalformedStatementError(
            f"{asset_class}: faltan columnas ESPECIE/Cantidad/Precio/Valor corriente"
        )

    positions: list[Position] = []
    for idx, row in enumerate(data_rows, start=1):
        ticker = _cell_str(row[c_ticker] if c_ticker < len(row) else "").upper()
        if not ticker or _norm_header(ticker) == "subtotal":
            continue
        quantity = _to_decimal(
            row[c_qty] if c_qty < len(row) else None,
            context=f"{asset_class}/{ticker}/cantidad",
        )
        price = _to_decimal(
            row[c_price] if c_price < len(row) else None,
            context=f"{asset_class}/{ticker}/precio",
        )
        total = _to_decimal(
            row[c_total] if c_total < len(row) else None,
            context=f"{asset_class}/{ticker}/total",
        )
        positions.append(
            Position(
                ticker=ticker,
                quantity=quantity,
                price=price,
                total=total,
                asset_class=asset_class,
            )
        )
    return positions


def _parse_totals_compact(rows: list[tuple[Any, ...]]) -> tuple[Decimal, Decimal]:
    total_ars: Decimal | None = None
    total_usd: Decimal | None = None
    for row in rows:
        key = _cell_str(row[0]).casefold()
        val = row[1] if len(row) > 1 else None
        if key in {"total ars", "total_ars", "ars"}:
            total_ars = _quantize_cent(_to_decimal(val, context="TOTALES/ARS"))
        elif key in {"total usd", "total_usd", "usd"}:
            total_usd = _quantize_cent(_to_decimal(val, context="TOTALES/USD"))
    if total_ars is None or total_usd is None:
        raise MalformedStatementError("TOTALES debe declarar 'Total ARS' y 'Total USD'")
    if total_usd == 0:
        raise MalformedStatementError("Total USD no puede ser cero (MEP indefinido)")
    return total_ars, total_usd


def _validate_portfolio_total_ars_compact(
    cash: list[CashBalance],
    positions: list[Position],
    total_ars: Decimal,
) -> None:
    cash_ars = next(c.amount for c in cash if c.currency == "ARS")
    positions_sum = sum((p.total for p in positions), Decimal("0.00"))
    computed = cash_ars + positions_sum
    if computed != total_ars:
        raise TotalsMismatchError(
            f"Total ARS declarado={total_ars} ≠ cash_ARS+posiciones={computed}"
        )


def _validate_portfolio_total_ars_broker(
    cash_valor_corriente: Decimal,
    positions: list[Position],
    total_ars: Decimal,
) -> None:
    """Total cartera bróker = suma VALOR CORRIENTE (cash convertido + posiciones)."""
    positions_sum = sum((p.total for p in positions), Decimal("0"))
    computed = cash_valor_corriente + positions_sum
    delta = abs(computed - total_ars)
    if delta > _BROKER_TOTAL_TOLERANCE:
        raise TotalsMismatchError(
            f"Total ARS declarado={total_ars} ≠ cash_valor+posiciones={computed} (delta={delta})"
        )


def parse_account_statement(path: str | Path) -> AccountSnapshot:
    """Parsea un .xlsx de estado de cuenta → AccountSnapshot scrubbeado.

    Detecta layout compact o broker_wide y normaliza al mismo contrato.
    Scrubbea titular/comitente → INV-001 antes de armar el snapshot.
    """
    xlsx_path = Path(path)
    if not xlsx_path.is_file():
        raise MalformedStatementError(f"Archivo inexistente: {xlsx_path}")

    rows = _read_sheet_rows(xlsx_path)
    if not rows:
        raise MalformedStatementError("El .xlsx está vacío")

    layout = _detect_layout(rows)
    if layout == "broker_wide":
        as_of, titular_raw, comitente_raw = _find_header_fields_broker(rows)
    else:
        as_of, titular_raw, comitente_raw = _find_header_fields_compact(rows)

    if titular_raw or comitente_raw:
        logger.info(
            "pii_scrubbed",
            extra={
                "investor_alias": INVESTOR_ALIAS,
                "had_titular": bool(titular_raw),
                "had_comitente": bool(comitente_raw),
            },
        )

    sections = _iter_sections(rows)
    for required in (SECTION_MONEDAS, *ASSET_SECTIONS):
        if required not in sections:
            raise MalformedStatementError(f"Falta la sección obligatoria '{required}'")

    cash_valor = Decimal("0")
    if layout == "broker_wide":
        cash, cash_valor = _parse_cash_broker(sections[SECTION_MONEDAS])
        positions: list[Position] = []
        for asset_class in ASSET_SECTIONS:
            positions.extend(_parse_positions_broker(sections[asset_class], asset_class))

        total_ars, total_usd = _find_header_totals_broker(rows)
        if total_ars is None or total_usd is None:
            if SECTION_TOTALES in sections:
                total_ars, total_usd = _parse_totals_compact(sections[SECTION_TOTALES])
            else:
                raise MalformedStatementError(
                    "Faltan totales de cabecera (TOTAL CARTERA / TOTAL USD) "
                    "y no hay sección TOTALES"
                )
        if total_usd == 0:
            raise MalformedStatementError("Total USD no puede ser cero (MEP indefinido)")
        _validate_portfolio_total_ars_broker(cash_valor, positions, total_ars)
    else:
        if SECTION_TOTALES not in sections:
            raise MalformedStatementError(f"Falta la sección obligatoria '{SECTION_TOTALES}'")
        cash = _parse_cash_compact(sections[SECTION_MONEDAS])
        positions = []
        for asset_class in ASSET_SECTIONS:
            positions.extend(_parse_positions_compact(sections[asset_class], asset_class))
        total_ars, total_usd = _parse_totals_compact(sections[SECTION_TOTALES])
        _validate_portfolio_total_ars_compact(cash, positions, total_ars)

    mep_implied = total_ars / total_usd  # Decimal exacto; sin round()

    snapshot = AccountSnapshot(
        investor_alias=INVESTOR_ALIAS,
        as_of=as_of,
        cash=cash,
        positions=positions,
        total_ars=total_ars,
        total_usd=total_usd,
        mep_implied=mep_implied,
    )

    logger.info(
        "statement_parsed",
        extra={
            "investor_alias": snapshot.investor_alias,
            "layout": layout,
            "positions": len(snapshot.positions),
            "total_ars": str(snapshot.total_ars),
            "total_usd": str(snapshot.total_usd),
            "mep_implied": str(snapshot.mep_implied),
        },
    )
    return snapshot
